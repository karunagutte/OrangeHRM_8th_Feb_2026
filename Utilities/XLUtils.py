import openpyxl

class Excel_methods:

   @staticmethod
   def count_rows(file_path,sheet_name):
       excel_file=openpyxl.load_workbook(file_path)
       sheet=excel_file[sheet_name]
       return sheet.max_row

   @staticmethod
   def read_data_from_excel_file(file_path,sheet_name,row_number,col_number):
       excel_file = openpyxl.load_workbook(file_path)
       sheet = excel_file[sheet_name]
       return sheet.cell(row=row_number,column=col_number).value

   @staticmethod
   def write_data_to_excel_file(file_path,sheet_name,row_number,col_number,data):
       excel_file = openpyxl.load_workbook(file_path)
       sheet = excel_file[sheet_name]
       sheet.cell(row_number,col_number).value=data
       excel_file.save(file_path)
       excel_file.close()